from flask import Flask, request
from flask import jsonify
import random
import openpyxl
import random

app = Flask(__name__)

    
@app.route("/students")
def students():
    students = {"elements": []}
    workbook = openpyxl.load_workbook('ANSC121.xlsx')
    sheet = workbook['Attendance']
    col = 'A'
    i = 2
    max_row = sheet.max_row # Get the maximum number of rows with data in the sheet
    for i in range(1, max_row + 1):
        cell_address = col + str(i)
        cell_value = sheet[cell_address].value
        if cell_value is not None:
            students["elements"].append(cell_value)
    print(max)
    return students


@app.route("/questions-and-choices")
def questions_with_choices():
    questions_with_choices = {"elements":[]}
    workbook = openpyxl.load_workbook('ANSC121.xlsx')
    sheet = workbook['Questions']
    max_row = sheet.max_row
    columns = ['A', 'B', 'C', 'D', 'E', 'F']
    #correct_answers = []

    for row in range(2, max_row + 1):
        choices = []
        question = ''
        for column in columns: 
            if column != 'A':       
                cell_address = sheet[column + str(row)] 
                #cellAddress_Color = cell_address.fill.start_color.index
                choice = cell_address.value
                choices.append(choice)
                # if cellAddress_Color == 'FFFFFF00':
                #     correct_answer = cell_address.coordinate
                #     correct_answers.append(correct_answer)
                #print(cell_address.coordinate, " : ",cellAddress_Color, " : ",choice)
            else:
                cell_address = sheet[column + str(row)]  
                question = cell_address.value 
        questions_with_choices["elements"].append({'question': question, 'choices': choices})
        #print(questions_with_choices)
    return jsonify(questions_with_choices)

@app.route('/submit-data', methods=['POST'])
def submit_data():
    
    data = request.json
    value = data.get('key') #this the answer from students
    print('this from the /submit-data endpoint: ', value)

    #extract the correct answers first before comparing the answer from students
    workbook = openpyxl.load_workbook('ANSC121.xlsx')
    sheet = workbook['Questions']
    max_row = sheet.max_row
    columns = ['B', 'C', 'D', 'E', 'F']
    correct_answers = []
    for row in range(2, max_row + 1):
        for column in columns: 
                cell_address = sheet[column + str(row)] 
                cellAddress_Color = cell_address.fill.start_color.index

                if cellAddress_Color == 'FFFFFF00':
                    correct_answer = cell_address.value
                    if correct_answer is not str:
                        correct_answer = str(correct_answer)
                    correct_answers.append(correct_answer)
    #calculate the score by comparing two arrays, value and correct_answers
    score = len(set(value) & set(correct_answers))
    #test_answers = set(value) & set(correct_answers)
    print("Student's Answers: ", value)
    print("Correct Answers:", correct_answers)
    print("Your Score: ", score)
    return jsonify(score)


    
if __name__ == "__main__":
    app.run(debug=True)
        