from flask import Flask, request
from flask import jsonify
import openpyxl
import string
import datetime

app = Flask(__name__)



def students():
    students = []
    workbook = openpyxl.load_workbook('ANSC121.xlsx')
    sheet = workbook['Attendance']
    col = 'A'
    i = 2
    max_row = sheet.max_row # Get the maximum number of rows with data in the sheet
    for i in range(2, max_row + 1):
        cell_address = col + str(i)
        cell_value = sheet[cell_address].value
        if cell_value is not None:
            students.append(cell_value)
    print(max)
    return students


@app.route("/questions-choices-loggers")
def questions_with_choices():
    current_date = datetime.datetime.now().strftime("%d/%m/%Y")
    elements = []
    workbook = openpyxl.load_workbook('ANSC121.xlsx')
    sheet = workbook['Questions']
    max_row = sheet.max_row
    columns = ['A', 'B', 'C', 'D', 'E', 'F']
    #correct_answers = []
    for row in range(2, max_row + 1):
        choices = []
        question = ''
        for column in columns: 
            if column != 'A':       
                cell_address = sheet[column + str(row)] 
                #cellAddress_Color = cell_address.fill.start_color.index
                choice = cell_address.value
                choices.append(choice)
                # if cellAddress_Color == 'FFFFFF00':
                #     correct_answer = cell_address.coordinate
                #     correct_answers.append(correct_answer)
                #print(cell_address.coordinate, " : ",cellAddress_Color, " : ",choice)
            else:
                cell_address = sheet[column + str(row)]  
                question = cell_address.value 
        elements.append({'question': question, 'choices': choices})
        #print(questions_with_choices)
        studyante = students()
        date_logger(current_date)
    return jsonify({"elements": elements, "studyante": studyante, "time": current_date})

def date_logger(current_date):
    
    workbook = openpyxl.load_workbook('ANSC121.xlsx')
    sheet = workbook['Attendance']
    already_logged = False
    columns = list(string.ascii_uppercase)[2:26]
    row = 1
    if not already_logged:
        for column in columns:
            cell_value = sheet[column + str(row)].value
            if cell_value is not None: 
                if cell_value == current_date:
                    already_logged = True
                    #print(current_date, f" has already been logged into {sheet[column + str(row)].coordinate}")
            elif cell_value is None and not already_logged:
                sheet[column + str(row)].value = current_date
                #print(current_date, f" is now being logged into {sheet[column + str(row)].coordinate}")
                already_logged = True
    workbook.save('ANSC121.xlsx')



@app.route('/submit-data', methods=['POST'])
def submit_data():
    current_date = datetime.datetime.now().strftime("%d/%m/%Y")
    data = request.json
    value = data.get('key') #this the answer from students
    student = str(data.get('student')) #this is the student number sent from the user
    #extract the correct answers first before comparing the answer from students
    workbook = openpyxl.load_workbook('ANSC121.xlsx')
    sheet = workbook['Questions']
    max_row = sheet.max_row
    columns = ['A', 'B', 'C', 'D', 'E', 'F']
    correct_answers = []
    for row in range(2, max_row + 1):
        for column in columns:
                if column == 'A':
                    question = sheet[column + str(row)].value 
                cell_address = sheet[column + str(row)] 
                cellAddress_Color = cell_address.fill.start_color.index

                if cellAddress_Color == 'FFFFFF00':
                    correct_answer = cell_address.value
                    if correct_answer is not str:
                        correct_answer = str(correct_answer)
                    correct_answers.append({question: correct_answer})
    #calculate the score by comparing two arrays, value and correct_answers
    #score = len(set(value) & set(correct_answers))
    #print('This is the score of the student', score)
    # print(f"STUDENT {student} SHOULD BE PRINTED HERE!")
    student_answer = []
    score = 0
    for question_and_answer in value:
        for correct_answer in correct_answers:
            if question_and_answer == correct_answer:
                score += 1    
    #print(student_answer)
    print(score)
    print('This is the score of the student', score)
    score_logger(score, student, current_date) 
    return jsonify(score)

def score_logger(score, student_number, current_date):
    columns = list(string.ascii_uppercase)[2:26]
    workbook = openpyxl.load_workbook('ANSC121.xlsx')
    sheet = workbook['Attendance']
    already_submitted = False
    maximum_row = sheet.max_row
    target_column = 'AA'
    target_row = 1
    target_coordinate = 'AA1'
    #first, find the column that contains the current date and grab that column
    #next is find the row that contains the students number
    #then, add the column and row, if that cell is empty, fill it with the score, else already submitted is True

    for column in columns:
        cell_value = sheet[column + str(1)].value
        if cell_value is not None:
            if cell_value == current_date:
                target_column = column
               # print(target_column, ' now find the row to complete the coordinates')
                for row in range(1, maximum_row + 1):
                    cell_value = str(sheet['A' + str(row)].value)
                    if cell_value is not None:
                        if cell_value == student_number:
                            target_row = row
                            target_coordinate = target_column + str(target_row)
                            if sheet[target_coordinate].value is not None:
                                already_submitted = True
    #print('in this case, the target coordinate is: ', target_coordinate)
    if not already_submitted:
        sheet[target_coordinate].value = score
    else:
        print(f"student {student_number} has already been submitted")


    workbook.save('ANSC121.xlsx')






    
if __name__ == "__main__":
    app.run(debug=True)
        